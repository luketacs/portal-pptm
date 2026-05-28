import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

ARQUIVO_MOVIMENTACOES = "Movimentações - 2026.xlsx"
ABA_MOVIMENTACOES     = "1-Movimentação dos produtos"
ARQUIVO_SAIDA         = "Entradas_15_Dias.xlsx"
JANELA_DIAS           = 15


# ---------------------------------------------------------------------------
# Leitura e filtragem
# ---------------------------------------------------------------------------

def carregar_entradas(caminho, data_corte):
    wb = openpyxl.load_workbook(caminho, read_only=True)

    if ABA_MOVIMENTACOES not in wb.sheetnames:
        print(f"ERRO: Aba '{ABA_MOVIMENTACOES}' não encontrada.")
        print(f"Abas disponíveis: {wb.sheetnames}")
        sys.exit(1)

    ws   = wb[ABA_MOVIMENTACOES]
    rows = list(ws.iter_rows(values_only=True))

    entradas = []
    for row in rows[1:]:
        codigo   = row[0]
        data     = row[10]
        quantidade = row[16]

        if codigo is None or data is None:
            continue
        if not isinstance(data, datetime):
            continue
        if data.date() < data_corte:
            continue

        try:
            qtd = float(quantidade) if quantidade not in (None, "", "|") else 0
        except (TypeError, ValueError):
            qtd = 0

        if qtd <= 0:
            continue

        custo_medio = row[5] or 0
        try:
            custo_medio = float(custo_medio)
        except (TypeError, ValueError):
            custo_medio = 0

        entradas.append({
            "data":        data,
            "codigo":      str(codigo),
            "descricao":   row[1] or "",
            "um":          row[2] or "",
            "grupo":       row[4] or "",
            "quantidade":  qtd,
            "custo_medio": custo_medio,
            "valor_total": qtd * custo_medio,
            "documento":   str(row[14] or ""),
            "referencia":  str(row[30] or "") if len(row) > 30 else "",
        })

    entradas.sort(key=lambda x: x["data"])
    return entradas


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------

def gerar_excel(entradas, data_corte, caminho_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Entradas {JANELA_DIAS} dias"

    hoje      = datetime.now().date()
    NUM_COLS  = 9

    # Estilos
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_fill   = PatternFill(start_color="243F60", end_color="243F60", fill_type="solid")
    alt_fill     = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill   = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Linha 1 — título
    ws.merge_cells(f"A1:{chr(64+NUM_COLS)}1")
    ws.row_dimensions[1].height = 30
    c           = ws["A1"]
    c.value     = (f"ENTRADAS NO ESTOQUE — ÚLTIMOS {JANELA_DIAS} DIAS  "
                   f"({data_corte.strftime('%d/%m/%Y')} a {hoje.strftime('%d/%m/%Y')})"
                   f"  —  Gerado em {datetime.now().strftime('%d/%m/%Y')}")
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = title_fill
    c.alignment = center

    # Linha 2 — cabeçalho
    headers = [
        "Nº", "Data", "Código", "Descrição", "UM", "Grupo",
        "Qtd. Entrada", "Custo Médio (R$)", "Valor Total (R$)",
    ]
    ws.row_dimensions[2].height = 30
    for col, h in enumerate(headers, 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = center
        c.border    = border

    # Linhas de dados
    for i, e in enumerate(entradas, 1):
        row_num = i + 2
        ws.row_dimensions[row_num].height = 18

        row_data = [
            i,
            e["data"].strftime("%d/%m/%Y"),
            e["codigo"],
            e["descricao"],
            e["um"],
            e["grupo"],
            e["quantidade"],
            e["custo_medio"],
            e["valor_total"],
        ]

        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col, val in enumerate(row_data, 1):
            c           = ws.cell(row=row_num, column=col, value=val)
            c.border    = border
            c.fill      = fill
            c.alignment = center if col in (1, 2, 5, 6, 7) else left
            if col in (8, 9):
                c.number_format = "R$ #,##0.00"
            elif col == 7:
                c.number_format = "#,##0.00"

    # Linha de totais
    total_row = len(entradas) + 3
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:F{total_row}")
    c       = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font  = Font(bold=True)
    c.fill  = total_fill
    c.alignment = center
    c.border    = border

    total_qtd = sum(e["quantidade"] for e in entradas)
    c = ws.cell(row=total_row, column=7, value=total_qtd)
    c.font = Font(bold=True); c.fill = total_fill
    c.number_format = "#,##0.00"; c.alignment = center; c.border = border

    total_val = sum(e["valor_total"] for e in entradas)
    c = ws.cell(row=total_row, column=9, value=total_val)
    c.font = Font(bold=True); c.fill = total_fill
    c.number_format = "R$ #,##0.00"; c.alignment = center; c.border = border

    # Resumo lateral
    resumo_col  = NUM_COLS + 2
    resumo_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    resumo_font = Font(bold=True, size=10)

    def resumo_linha(row, label, valor, fmt=None):
        ws.row_dimensions[row].height = 18
        cl = ws.cell(row=row, column=resumo_col, value=label)
        cl.font = resumo_font; cl.fill = resumo_fill; cl.border = border
        cl.alignment = left

        cv = ws.cell(row=row, column=resumo_col + 1, value=valor)
        cv.fill = resumo_fill; cv.border = border; cv.alignment = center
        if fmt:
            cv.number_format = fmt

    resumo_linha(1, "Período",           f"{data_corte.strftime('%d/%m/%Y')} → {hoje.strftime('%d/%m/%Y')}")
    resumo_linha(2, "Total de lançamentos", len(entradas))
    resumo_linha(3, "Materiais distintos",
                 len({e["codigo"] for e in entradas}))
    resumo_linha(4, "Qtd. total entrada", total_qtd, "#,##0.00")
    resumo_linha(5, "Valor total (R$)",   total_val,  "R$ #,##0.00")

    ws.column_dimensions[chr(64 + resumo_col)].width     = 22
    ws.column_dimensions[chr(64 + resumo_col + 1)].width = 18

    # Larguras das colunas principais
    col_widths = {
        "A": 5, "B": 12, "C": 13, "D": 50,
        "E": 6, "F": 8,  "G": 13, "H": 18, "I": 18,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"
    wb.save(caminho_saida)
    return total_val, total_qtd


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    pasta       = os.path.dirname(os.path.abspath(__file__))
    entrada     = os.path.join(pasta, ARQUIVO_MOVIMENTACOES)
    saida       = os.path.join(pasta, ARQUIVO_SAIDA)
    data_corte  = datetime.now().date() - timedelta(days=JANELA_DIAS)

    if not os.path.exists(entrada):
        print(f"ERRO: '{ARQUIVO_MOVIMENTACOES}' não encontrado em:\n  {pasta}")
        sys.exit(1)

    print(f"Lendo movimentações (últimos {JANELA_DIAS} dias: a partir de {data_corte.strftime('%d/%m/%Y')})...")
    entradas = carregar_entradas(entrada, data_corte)
    print(f"  Lançamentos de entrada encontrados: {len(entradas)}")
    print(f"  Materiais distintos: {len({e['codigo'] for e in entradas})}")

    if not entradas:
        print("  Nenhuma entrada no período. Relatório não gerado.")
        sys.exit(0)

    try:
        total_val, total_qtd = gerar_excel(entradas, data_corte, saida)
    except PermissionError:
        print(f"\nERRO: Feche o arquivo '{ARQUIVO_SAIDA}' no Excel e execute novamente.")
        sys.exit(1)

    print(f"  Qtd. total entrada: {total_qtd:,.0f}")
    print(f"  Valor total:        R$ {total_val:,.2f}")
    print(f"\nRelatório salvo em:\n  {saida}")


if __name__ == "__main__":
    main()
