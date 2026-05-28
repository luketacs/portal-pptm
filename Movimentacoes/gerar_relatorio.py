import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

ARQUIVO_MOVIMENTACOES = "Movimentações - 2026.xlsx"
ARQUIVO_SOLICITACOES  = "Solicitacoes.xlsx"
ARQUIVO_RELATORIO_ARY = "Relatorio Ary.xlsx"
ARQUIVO_SAIDA         = "Materiais_Aguardando_Retirada_2026.xlsx"
ABA_MOVIMENTACOES     = "1-Movimentação dos produtos"
ABA_SOLICITACOES      = "Listagem do Browse"


# ---------------------------------------------------------------------------
# Leitura
# ---------------------------------------------------------------------------

def carregar_planilha(nome_arquivo, nome_aba=None):
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_arquivo)
    if not os.path.exists(caminho):
        print(f"ERRO: '{nome_arquivo}' não encontrado em:\n  {caminho}")
        sys.exit(1)
    wb = openpyxl.load_workbook(caminho, read_only=True)
    if nome_aba is None:
        nome_aba = wb.sheetnames[0]
    elif nome_aba not in wb.sheetnames:
        print(f"ERRO: Aba '{nome_aba}' não encontrada em '{nome_arquivo}'.")
        print(f"Abas disponíveis: {wb.sheetnames}")
        sys.exit(1)
    return list(wb[nome_aba].iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# Análise de movimentações → materiais com entrada e sem saída
# ---------------------------------------------------------------------------

def analisar_movimentacoes(rows):
    materiais = defaultdict(lambda: {
        "descricao":      "",
        "um":             "",
        "grupo":          "",
        "custo_medio":    0,
        "saldo_qtd":      0,
        "total_entradas": 0,
        "total_saidas":   0,
        "ultima_data":    None,
    })

    def to_float(val):
        try:
            return float(val) if val not in (None, "", "|") else 0
        except (TypeError, ValueError):
            return 0

    for row in rows[1:]:
        codigo = row[0]
        if codigo is None:
            continue
        m = materiais[codigo]
        m["descricao"]      = row[1] or ""
        m["um"]             = row[2] or ""
        m["grupo"]          = row[4] or ""
        m["custo_medio"]    = row[5] or 0
        m["saldo_qtd"]      = row[6] or 0
        m["total_entradas"] += to_float(row[16])
        m["total_saidas"]   += to_float(row[22])
        if row[10] and (m["ultima_data"] is None or row[10] > m["ultima_data"]):
            m["ultima_data"] = row[10]

    return {
        str(k): v
        for k, v in materiais.items()
        if v["total_entradas"] > 0 and v["total_saidas"] == 0
    }


# ---------------------------------------------------------------------------
# Cruzamento com Solicitações → SA, Ordem, Recebedor por produto
# ---------------------------------------------------------------------------

def cruzar_solicitacoes(rows_sol, codigos_sem_saida):
    """
    Retorna dict: codigo -> lista de dict {sa, ordem, recebedores, qtd_sol, qtd_atend}
    Cada item é um par único (produto, SA).
    """
    por_produto_sa = defaultdict(lambda: {
        "ordens":      set(),
        "recebedores": set(),
        "qtd_sol":     0,
        "qtd_atend":   0,
    })

    for row in rows_sol[2:]:
        produto = str(row[2] or "").strip()
        if produto not in codigos_sem_saida:
            continue

        sa        = str(row[0] or "").strip()
        ord_prod  = str(row[7] or "").strip()
        ordem_6   = ord_prod[:6] if ord_prod else ""
        recebedor = str(row[12] or "").strip()
        qtd_sol   = float(row[5] or 0)
        qtd_atend = float(row[11] or 0)

        key = (produto, sa)
        if ordem_6:
            por_produto_sa[key]["ordens"].add(ordem_6)
        if recebedor:
            por_produto_sa[key]["recebedores"].add(recebedor)
        por_produto_sa[key]["qtd_sol"]   += qtd_sol
        por_produto_sa[key]["qtd_atend"] += qtd_atend

    # Reorganizar: codigo -> [{sa, ordens, recebedores, ...}]
    resultado = defaultdict(list)
    for (produto, sa), info in por_produto_sa.items():
        resultado[produto].append({
            "sa":          sa,
            "ordens":      sorted(info["ordens"]),
            "recebedores": sorted(info["recebedores"]),
            "qtd_sol":     info["qtd_sol"],
            "qtd_atend":   info["qtd_atend"],
        })

    # Ordenar cada lista de SAs por número de SA
    for lst in resultado.values():
        lst.sort(key=lambda x: x["sa"])

    return resultado


# ---------------------------------------------------------------------------
# SAs encerradas — Relatorio Ary
# ---------------------------------------------------------------------------

def carregar_sas_encerradas(rows_ary):
    """
    Retorna conjunto de tuplas (sa_6digitos, produto_str) cujos TODOS os itens
    no Relatorio Ary têm Status da SA (col 17) = 'E' (encerrada).
    """
    pares = defaultdict(set)
    for row in rows_ary[1:]:
        sa      = str(row[1]).zfill(6)
        produto = str(row[3])
        status  = row[17]
        pares[(sa, produto)].add(status)

    return {par for par, statuses in pares.items() if all(s == "E" for s in statuses)}


# ---------------------------------------------------------------------------
# Montagem das linhas do relatório
# ---------------------------------------------------------------------------

def filtrar_sas_por_estoque(sas, disponivel):
    """
    Recebe lista de SAs já ordenadas (mais antiga primeiro) e o estoque disponível.
    Retorna apenas as SAs que o estoque ainda consegue cobrir (acumulando qtd_sol).
    Cada SA recebe 'qtd_atende' indicando quanto do estoque ela efetivamente recebe.
    Se nenhuma couber integralmente, retorna ao menos a mais antiga (prioridade de fila).
    """
    incluidas = []
    acumulado = 0.0
    for sa_info in sas:
        if acumulado >= disponivel:
            break
        restante   = disponivel - acumulado
        qtd_atende = min(sa_info["qtd_sol"], restante)
        incluidas.append({**sa_info, "qtd_atende": qtd_atende})
        acumulado += sa_info["qtd_sol"]
    if not incluidas and sas:
        sa = sas[0]
        incluidas = [{**sa, "qtd_atende": min(sa["qtd_sol"], disponivel)}]
    return incluidas


def montar_linhas(sem_saida, solicitacoes_por_produto, sas_encerradas=None):
    """
    Retorna lista de grupos, cada grupo = lista de linhas de um material.
    Primeira linha do grupo é o item principal (is_subitem=False).
    Linhas seguintes são sub-itens (is_subitem=True), com colunas de material vazias.
    SAs encerradas (todas as SA+produto com Status da SA = 'E') são excluídas.
    SAs que excedem o estoque disponível também são excluídas (ficam pendentes).
    """
    if sas_encerradas is None:
        sas_encerradas = set()

    grupos = []

    for codigo, info in sem_saida.items():
        sas_raw = solicitacoes_por_produto.get(codigo, [])
        # Remove SAs encerradas antes de qualquer outro filtro
        sas = [s for s in sas_raw if (s["sa"], codigo) not in sas_encerradas]

        base = {
            "codigo":      codigo,
            "descricao":   info["descricao"],
            "um":          info["um"],
            "grupo":       info["grupo"],
            "entradas":    info["total_entradas"],
            "saldo":       info["saldo_qtd"],
            "custo_medio": info["custo_medio"],
            "valor_total": info["total_entradas"] * info["custo_medio"],
            "ultima_data": info["ultima_data"],
        }

        if not sas:
            grupos.append([{
                **base,
                "sa": "", "ordens": [], "recebedores": [],
                "qtd_sol": 0, "qtd_atende": 0,
                "is_subitem": False,
                "secao": "sem_sa",
            }])
        else:
            sas_validas = filtrar_sas_por_estoque(sas, info["total_entradas"])
            grupo = []
            for i, sa_info in enumerate(sas_validas):
                linha = {**sa_info, "is_subitem": i > 0, "secao": "com_sa"}
                if i == 0:
                    linha = {**base, **linha}
                else:
                    linha["codigo"]      = codigo
                    linha["ultima_data"] = info["ultima_data"]
                grupo.append(linha)
            grupos.append(grupo)

    # Ordenar grupos pela data da última movimentação do material
    grupos.sort(key=lambda g: g[0].get("ultima_data") or datetime.min)

    # Achatar para lista plana preservando a ordem dos sub-itens dentro do grupo
    return [linha for grupo in grupos for linha in grupo]


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------

def gerar_excel(linhas, caminho_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiais sem Saída"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_fill  = PatternFill(start_color="243F60", end_color="243F60", fill_type="solid")
    alt_fill    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    sem_sa_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    NUM_COLS = 15

    # Linha 1 – título
    ws.merge_cells(f"A1:{chr(64+NUM_COLS)}1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value     = f"MATERIAIS COM ENTRADA SEM SAÍDA DO ALMOXARIFADO  —  Gerado em {datetime.now().strftime('%d/%m/%Y')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = title_fill
    c.alignment = center

    # Linha 2 – cabeçalho
    headers = [
        "Nº", "Código", "Descrição", "UM", "Grupo",
        "Qtd. Entrada", "Saldo Qtd.", "Custo Médio (R$)", "Valor Total (R$)",
        "Últ. Moviment.", "Nr. SA", "Ordem", "Qtd. Solicitada", "Qtd. Atende", "Recebedor(es)",
    ]
    ws.row_dimensions[2].height = 35
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = center
        c.border    = border

    # Estilos extras
    subitem_fill_alt   = PatternFill(start_color="ECF0FA", end_color="ECF0FA", fill_type="solid")
    subitem_fill_plain = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    parcial_fill       = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    secao1_fill        = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    secao2_fill        = PatternFill(start_color="538135", end_color="538135", fill_type="solid")
    subtotal_fill      = PatternFill(start_color="D6E4BC", end_color="D6E4BC", fill_type="solid")
    dashed_side        = Side(style="dashed", color="AAAAAA")
    border_sub         = Border(left=thin, right=thin, top=dashed_side, bottom=dashed_side)
    gray_font          = Font(color="888888", size=10)

    # Separar as duas seções
    com_sa = [l for l in linhas if l.get("secao") == "com_sa"]
    sem_sa = [l for l in linhas if l.get("secao") == "sem_sa"]

    # -----------------------------------------------------------------------
    # Função auxiliar: escreve um bloco de linhas e retorna (row_num, codigos_vistos_bloco)
    # -----------------------------------------------------------------------
    def escrever_bloco(bloco, row_start, n_item_start, fill_sub_ref):
        n_item       = n_item_start
        row_num      = row_start
        codigos_vis  = {}
        fill_sub     = fill_sub_ref   # atualizado a cada item principal

        for linha in bloco:
            is_sub = linha.get("is_subitem", False)
            row_num += 1
            ws.row_dimensions[row_num].height = 18 if is_sub else 20

            sa     = linha.get("sa", "")
            ordens = ", ".join(linha.get("ordens", []))
            receb  = ", ".join(linha.get("recebedores", []))

            if not is_sub:
                n_item  += 1
                usa_alt  = (n_item % 2 == 0)
                fill_pri = alt_fill          if usa_alt else PatternFill()
                fill_sub = subitem_fill_alt  if usa_alt else subitem_fill_plain

                data_fmt = linha["ultima_data"].strftime("%d/%m/%Y") if linha.get("ultima_data") else ""
                row_data = [
                    n_item,
                    linha["codigo"],
                    linha["descricao"],
                    linha["um"],
                    linha["grupo"],
                    linha["entradas"],
                    linha["saldo"],
                    linha["custo_medio"],
                    linha["valor_total"],
                    data_fmt,
                    sa,
                    ordens,
                    linha.get("qtd_sol", 0) or "",
                    linha.get("qtd_atende", 0) or "",
                    receb,
                ]
                fill_uso = sem_sa_fill if not sa else fill_pri

                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.border    = border
                    c.fill      = fill_uso
                    c.alignment = center if col in (1, 4, 5, 6, 7, 10, 11, 12, 13, 14) else left
                    if col == 2:
                        c.font = Font(bold=True)
                    if col in (8, 9):
                        c.number_format = "R$ #,##0.00"
                    elif col in (6, 7, 13, 14):
                        c.number_format = "#,##0.00"
                    if col == 14:
                        qtd_s = linha.get("qtd_sol", 0) or 0
                        qtd_a = linha.get("qtd_atende", 0) or 0
                        if qtd_s > 0 and qtd_a < qtd_s:
                            c.fill = parcial_fill

                if linha["codigo"] not in codigos_vis:
                    codigos_vis[linha["codigo"]] = linha.get("valor_total", 0)

            else:
                for col in range(1, 11):
                    c = ws.cell(row=row_num, column=col, value="")
                    c.fill   = fill_sub
                    c.border = Border(
                        left=thin if col == 1 else Side(),
                        right=thin if col == 10 else Side(),
                        top=dashed_side, bottom=dashed_side,
                    )
                qtd_s = linha.get("qtd_sol", 0) or 0
                qtd_a = linha.get("qtd_atende", 0) or 0
                for col, val in zip([11, 12, 13, 14, 15], [sa, ordens, qtd_s or "", qtd_a or "", receb]):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.fill      = fill_sub
                    c.font      = gray_font
                    c.border    = border_sub
                    c.alignment = center if col in (11, 12, 13, 14) else left
                    if col in (13, 14):
                        c.number_format = "#,##0.00"
                    if col == 14 and qtd_s > 0 and qtd_a < qtd_s:
                        c.fill = parcial_fill

        return row_num, n_item, codigos_vis

    def escrever_cabecalho_secao(row_num, texto, fill_secao, qtd_itens):
        row_num += 1
        ws.row_dimensions[row_num].height = 22
        ws.merge_cells(f"A{row_num}:{chr(64+NUM_COLS)}{row_num}")
        c = ws.cell(row=row_num, column=1,
                    value=f"  {texto}  —  {qtd_itens} material(is)")
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = fill_secao
        c.alignment = left
        return row_num

    def escrever_subtotal(row_num, bloco_codigos):
        row_num += 1
        ws.row_dimensions[row_num].height = 18
        ws.merge_cells(f"A{row_num}:E{row_num}")
        c = ws.cell(row=row_num, column=1, value="Subtotal")
        c.font = Font(bold=True, italic=True); c.alignment = center; c.border = border; c.fill = subtotal_fill

        qtd = sum(l["entradas"] for l in linhas if not l.get("is_subitem") and l["codigo"] in bloco_codigos)
        c = ws.cell(row=row_num, column=6, value=qtd)
        c.font = Font(bold=True, italic=True); c.fill = subtotal_fill
        c.number_format = "#,##0.00"; c.alignment = center; c.border = border

        val = sum(bloco_codigos.values())
        c = ws.cell(row=row_num, column=9, value=val)
        c.font = Font(bold=True, italic=True); c.fill = subtotal_fill
        c.number_format = "R$ #,##0.00"; c.alignment = center; c.border = border
        return row_num, val

    # -----------------------------------------------------------------------
    # Renderizar seção 1 — Com SA
    # -----------------------------------------------------------------------
    n_materiais_com_sa = sum(1 for l in com_sa if not l.get("is_subitem"))
    row_num = escrever_cabecalho_secao(2, "SEÇÃO 1 — MATERIAIS COM SA PENDENTE DE RETIRADA",
                                       secao1_fill, n_materiais_com_sa)
    row_num, n_item, cod_com_sa = escrever_bloco(com_sa, row_num, 0, subitem_fill_plain)
    row_num, _ = escrever_subtotal(row_num, cod_com_sa)

    # -----------------------------------------------------------------------
    # Renderizar seção 2 — Sem SA
    # -----------------------------------------------------------------------
    n_materiais_sem_sa = len(sem_sa)
    row_num = escrever_cabecalho_secao(row_num + 1, "SEÇÃO 2 — MATERIAIS SEM SA CORRESPONDENTE",
                                       secao2_fill, n_materiais_sem_sa)
    row_num, n_item, cod_sem_sa = escrever_bloco(sem_sa, row_num, n_item, subitem_fill_plain)
    row_num, _ = escrever_subtotal(row_num, cod_sem_sa)

    codigos_vistos = {**cod_com_sa, **cod_sem_sa}

    # -----------------------------------------------------------------------
    # Total geral
    # -----------------------------------------------------------------------
    total_row = row_num + 1
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:E{total_row}")
    c = ws.cell(row=total_row, column=1, value="TOTAL GERAL")
    c.font = Font(bold=True); c.alignment = center; c.border = border; c.fill = total_fill

    total_qtd = sum(l["entradas"] for l in linhas if not l.get("is_subitem"))
    c = ws.cell(row=total_row, column=6, value=total_qtd)
    c.font = Font(bold=True); c.fill = total_fill
    c.number_format = "#,##0.00"; c.alignment = center; c.border = border

    total_val = sum(codigos_vistos.values())
    c = ws.cell(row=total_row, column=9, value=total_val)
    c.font = Font(bold=True); c.fill = total_fill
    c.number_format = "R$ #,##0.00"; c.alignment = center; c.border = border

    # Legenda
    leg_row = total_row + 2
    ws.cell(row=leg_row, column=1, value="Legenda:").font = Font(bold=True)
    c2 = ws.cell(row=leg_row, column=4, value="Sub-item (SA adicional do mesmo material)")
    c2.fill = subitem_fill_alt; c2.border = border; c2.font = gray_font
    c3 = ws.cell(row=leg_row, column=7, value="Atendimento parcial (estoque insuficiente para a qtd. solicitada)")
    c3.fill = parcial_fill; c3.border = border

    # Larguras
    col_widths = {"A":5,"B":13,"C":50,"D":6,"E":8,"F":13,"G":11,"H":18,"I":18,"J":14,"K":10,"L":10,"M":14,"N":12,"O":35}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"
    wb.save(caminho_saida)
    return total_val


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    pasta = os.path.dirname(os.path.abspath(__file__))

    print("Lendo movimentações...")
    rows_mov = carregar_planilha(ARQUIVO_MOVIMENTACOES, ABA_MOVIMENTACOES)
    sem_saida = analisar_movimentacoes(rows_mov)
    print(f"  Materiais com entrada sem saída: {len(sem_saida)}")

    print("Lendo solicitações...")
    rows_sol = carregar_planilha(ARQUIVO_SOLICITACOES, ABA_SOLICITACOES)
    sol_por_produto = cruzar_solicitacoes(rows_sol, set(sem_saida.keys()))
    com_match = sum(1 for c in sem_saida if c in sol_por_produto)
    print(f"  Materiais com SA encontrada: {com_match} | Sem correspondência: {len(sem_saida) - com_match}")

    print("Lendo Relatório Ary (SAs encerradas)...")
    rows_ary = carregar_planilha(ARQUIVO_RELATORIO_ARY, None)
    sas_encerradas = carregar_sas_encerradas(rows_ary)
    print(f"  Pares (SA, Produto) encerrados: {len(sas_encerradas)}")

    linhas = montar_linhas(sem_saida, sol_por_produto, sas_encerradas)
    print(f"  Total de linhas no relatório: {len(linhas)}")

    saida = os.path.join(pasta, ARQUIVO_SAIDA)
    try:
        total_val = gerar_excel(linhas, saida)
    except PermissionError:
        print(f"\nERRO: Feche o arquivo '{ARQUIVO_SAIDA}' no Excel e execute novamente.")
        sys.exit(1)

    print(f"  Valor total estimado em estoque: R$ {total_val:,.2f}")
    print(f"\nRelatório salvo em:\n  {saida}")


if __name__ == "__main__":
    main()
